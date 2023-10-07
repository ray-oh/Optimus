#!/usr/bin/env python
# coding: utf-8

"""
Module:         generalAutomationScript.py
Description:    RPA automation with Excel front end
Created:        30 Jul 2022

Versions:
20210216    Refactor KB Quest code - reusable sub routines - sub_TXT_*
            helper functions - hoverClick, hoverRclick, waitImage, waitImageDisappear, try_catch
            Logging
            Reorganize assets - image, log, output folders

"""
from prefect import task, flow, get_run_logger, context
from prefect.task_runners import SequentialTaskRunner

from config import variables, constants, STARTFILE, CWD_DIR, IMAGE_DIR, yesterdayYYYYMMDD, log_space, RPABROWSER
from pathlib import Path, PureWindowsPath
import rpa as r
from browser import Browser
p = Browser()

from auto_helper_lib import dfObjList, dfKey_value, readExcelConfig, updateConstants, CriticalAccessFailure
from auto_utility_PDF_Image import cropImage, createPDFfromImages, addContentPDF
from auto_utility_file import GetRecentCreatedFile, runInBackground, renameFile
from auto_utility_browser import chromeZoom, waitIdentifierExist, waitIdentifierDisappear, exitProg, snapImage
from auto_utility_parsers import parseArguments, regexSearch
from auto_utility_email import EmailsSender

def _otherRunCode(df, code, codeID, codeValue, objVar):

    # code for handling block codes - extract parameters if that is defined e.g. command(parameter)
    # parameters are in json string format e.g. {"name":"value","name":"value"}
    import re  
    param = re.search('\((.*?)\)', code, re.IGNORECASE)
    if param and code.strip()[-1:]==')':               # if parameters exist
        parameters = param.group(1).strip()
        command = re.search('(.*?)\(', code, re.IGNORECASE)        
        code = command.group(1).strip()
        print(code, parameters)

        import json
        #jsonString = '{"file":"C:\\Users\\roh\\Downloads\\d5c7a4f7-b9a7-4d1e-904e-ce7349e0f27c.xlsx", "country": "All"}'
        #jsonString = '{"file": "C:/Users/roh/Downloads/d5c7a4f7-b9a7-4d1e-904e-ce7349e0f27c.xlsx", "country": "All"}'
        variables.update(json.loads(parameters))
        print(variables)

    if False: pass
    # ------------------ code processing functions ------------------------------
    elif codeID.lower() == 'rem'.lower():   pass                    # remarks - do nothing
    elif codeID.lower() == 'print'.lower():  _print(codeValue)
    elif codeID.lower() == 'log':  _log(codeValue)
    elif codeID.lower() == 'exit'.lower():    _exit()
    elif codeID.lower() == 'exitError'.lower():    _exitError(codeValue)                  # exit with an error code e.g. exitError:2
    elif codeID.lower() == 'raiseError'.lower():    _raiseError(codeValue)                # exit with an error code e.g. exitError:2
    elif codeID.lower() == 'if'.lower(): return _if(codeValue, df, objVar)                # codeValue = condition : if true block, if false block or if empty then pass

    elif code in df[(df.Type == 'list')]['Object'].dropna().values.tolist(): return _isCodeList(df, code, objVar)          #run Block of Code
    elif codeID.lower() == 'runTask'.lower() and codeValue in df[(df.Type == 'list')]['Object'].dropna().values.tolist(): return _isCodeListTask(df, codeValue, objVar)          #run Block of Code    .with_options(name=codeValue.split(',')[0].strip())    
    elif codeID.lower() == 'runFlow'.lower() and codeValue in df[(df.Type == 'list')]['Object'].dropna().values.tolist(): return _isCodeListFlow(df, codeValue, objVar)          #run Block of Code        
    elif codeID.lower() == 'runModule'.lower(): return _runModule(codeValue, df, objVar)                  #runModule:sheet, codeblock, excelfile
    elif codeID.lower() == 'runTask'.lower(): return _runTask(codeValue, df, objVar)                  #runModule:sheet, excelfile .with_options(name=codeValue.split(',')[0].strip())
    elif codeID.lower() == 'runFlow'.lower(): return _runFlow(codeValue, df, objVar)                  #runModule:sheet, excelfile .with_options(name=codeValue.split(',')[0].strip())
    elif codeID.lower() == 'codeList'.lower(): return _codeList(codeValue, df, objVar)
    elif codeID.lower() == 'wait'.lower():  return _wait(codeValue, df, objVar)                        # wait:time_sec,identifier,run_code
    elif codeID.lower() == 'waitDisappear'.lower(): return _waitDisappear(codeValue, df, objVar)       # waitDisappear:time_sec,identifier,run_code
    elif codeID.lower() == 'waitFile'.lower(): return _waitFile(codeValue, df, objVar)                  # waitFile: file_pattern, action, timeout, actionIfTimeout
    elif codeID.lower() == 'iterate'.lower(): return _iterate(codeValue, df, objVar)                    # iterate: objlists, runCodelist e.g. iterate: @url_pages : openPage
    elif codeID.lower() == 'iterationCount'.lower(): _iterationCount(codeValue, df, objVar)
    elif codeID.lower() == 'loopWhile'.lower(): return _loopWhile(codeValue, df, objVar)                # while:condition:do this:repeat every X sec
    elif codeID.lower() == 'test'.lower(): _test(codeValue, df, objVar)

    # ------------------ Custom functions ------------------------------
    elif codeID.lower() == 'regexSearch'.lower():        _regexSearch(codeValue)                           # regexSearch:strPattern, strSearch, variable_name e.g. regexSearch:PACIFIC.ASIA.(..........),Last data was imported at,lastDataUpdate
    elif codeID.lower() == 'createPDF'.lower():   _createPDF(codeValue, df)

    #add_page_numbers(saveFile, pageTitles)
    #addContentPDF(pdf_path, pageTitles, file_extension = '_' + yesterdayYYYYMMDD + '.pdf')
    #addContentPDF:pageTitles,sourcePDF,targetPDF
    elif codeID.lower() == 'addContentPDF'.lower():  _addContentPDF(codeValue, df)
    elif codeID.lower() == 'cropImage'.lower():     _cropImage(codeValue, df)                              # cropImage:files, savefiles, left, top, right, bottom, boolPercentage = True/False
    elif codeID.lower() == 'runExcelMacro'.lower():       _runExcelMacro(codeValue)                          # runExcelMacro:excel, macro
    elif codeID.lower() == 'runPowerShellScript'.lower():       _runPowerShellScript(codeValue)                          # runExcelMacro:excel, macro
    elif codeID.lower() == 'runBatchScript'.lower():       _runBatchScript(codeValue)                          # runExcelMacro:excel, macro
    elif codeID.lower() == 'triggerRPAScript'.lower():       _triggerRPAScript(codeValue)                          # triggerRPAScript:command with flags    
    elif codeID.lower() == 'openProgram'.lower():       _openProgram(codeValue)                          # openProgram:path
    elif codeID.lower() == 'focusWindow'.lower():       _focusWindow(codeValue)                          # focusWindow:windowName
    elif codeID.lower() == 'runJupyterNb'.lower():       _runJupyterNb(codeValue)                          # runJupyterNb:notebook_file, parameters
    elif codeID.lower() == 'mergeFiles'.lower():    _mergeFiles(codeValue, df)                               # mergeFiles: fileList, keep, uniqueColumnsList, fileName, encoding 
    elif codeID.lower() == 'dropNRowsExcel'.lower():   _dropNRowsExcel(codeValue)                                # merge files.  merge:newFile,oldFile,olderFile
    elif codeID.lower() == 'DFpromoteHeader'.lower(): _DFpromoteHeader(codeValue) # merge files.  merge:newFile,oldFile,olderFile
    elif codeID.lower() == 'DFsaveToExcel'.lower():   _DFsaveToExcel(codeValue)  # merge files.  merge:newFile,oldFile,olderFile
    elif codeID.lower() == 'DFreadExcel'.lower():     _DFreadExcel(codeValue) # read Excel to dataframe.  DFreadExcel:filename, variableDataFrameName
    elif codeID.lower() == 'DFsort'.lower():      _DFsort(codeValue) # merge files.  merge:newFile,oldFile,olderFile
    elif codeID.lower() == 'DFdropDuplicates'.lower():    _DFdropDuplicates(codeValue) # merge files.  merge:newFile,oldFile,olderFile
    elif codeID.lower() == 'DFconcatenate'.lower():    _DFconcatenate(codeValue)    # merge files.  merge:newFile,oldFile,olderFile
    elif codeID.lower() == 'DFcreate'.lower(): _DFcreate(codeValue)
    # ------------------ Custom automation functions using pywinauto ------------------------------
    #elif codeID.lower() == 'openProgram'.lower():       _openProgram(codeValue)                          # keyboard_pwa:path    
    elif codeID.lower() == 'keyboard_pwa'.lower():  _keyboard_pwa(codeValue)        # key press e.g. [home] [end] [insert] [f1] .. [f15] [shift] [ctrl] [alt] [win] [cmd] [enter] [space] [tab] [esc] [backspace] [delete] [clear]

    # ------------------ Browser / Windows functions ------------------------------
    elif codeID.lower() == 'chromeZoom'.lower():  _chromeZoom(codeValue)
    elif codeID.lower() == 'copyFile'.lower():   _copyFile(codeValue)       # copyFile:source,dest  e.g. copy file to one drive sync folder
    elif codeID.lower() == 'moveFile'.lower():    _moveFile(codeValue)    # moveFile:source,dest  e.g. copy file to one drive sync folder
    elif codeID.lower() == 'makeDir'.lower():     _makeDir(codeValue)     # makeDir:pathname
    elif codeID.lower() == 'removeFile'.lower():  _removeFile(codeValue)        # removeFile:source,dest  e.g. copy file to one drive sync folder
    elif codeID.lower() == 'renameRecentDownloadFile'.lower(): _renameRecentDownloadFile(codeValue) # renameRecentDownloadFile:saveName,path,fileExtension,withinLastSec
    elif codeID.lower() == 'touchFile'.lower(): _touchFile(codeValue) # touch file https://stackoverflow.com/questions/1158076/implement-touch-using-python
    elif codeID.lower() == 'raiseEvent'.lower():   _raiseEvent(codeValue)        # raise event
    elif codeID.lower() == 'ifEvent'.lower():  return _ifEvent(codeValue, df, objVar)  # if event do action
    elif codeID.lower() == 'ifTest'.lower():  return _ifTest(codeValue, df, objVar)  # if event do action

    # ------------------ RPA functions ------------------------------
    elif codeID.lower() == 'runInBackground'.lower():   _runInBackground()  # run automation in background mode without user attendance
    elif codeID.lower() == 'initializeRPA'.lower(): _initializeRPA(codeValue)
    elif codeID.lower() == 'closeRPA'.lower():  _closeRPA()
    elif codeID.lower() == 'url'.lower():       _url(codeValue, df)         # url:OKTA or url:<URL_Dclick_Pages:key> or url:@<URL_Dclick_Pages:@columnHeader>
    elif codeID.lower() == 'urls'.lower():   _urls(codeValue, objVar)        # No longer required - can remove
    elif codeID.lower() == 'read'.lower():      _read(codeValue)        # read:checkUserName=okta-signin-username
    elif codeID.lower() == 'checkVariable'.lower(): _checkVariable(codeValue)    # checkVariable:checkUserName
    elif codeID.lower() == 'set'.lower():       _set(codeValue)             # set:checkUserName=value
    elif codeID.lower() == 'increment'.lower():       _increment(codeValue)             # increment:counter, 1    
    elif codeID.lower() == 'urlcontains'.lower(): _urlcontains(codeValue)   # urlcontains:value_to_search,variable_result_true_false
    elif codeID.lower() == 'keyboard'.lower():  _keyboard(codeValue)        # key press e.g. [home] [end] [insert] [f1] .. [f15] [shift] [ctrl] [alt] [win] [cmd] [enter] [space] [tab] [esc] [backspace] [delete] [clear]
    elif codeID.lower() == 'type'.lower():      _type(codeValue)            # type
    elif codeID.lower() == 'rclick'.lower():    _rclick(codeValue)          # right click
    elif codeID.lower() == 'present'.lower():   _present(codeID, codeValue) # right click
    elif codeID.lower() == 'exist'.lower():     _exist(codeID, codeValue)   # Waits until the timeout for an element to exist and returns a JavaScript true or false
    elif codeID.lower() == 'focus'.lower():   _focus(codeID, codeValue) # focus() - app_to_focus (full name of app) - make application in focus
    elif codeID.lower() == 'popup'.lower():   _popup(codeID, codeValue) # popup(url) - focus tagui rpa on specific tab/browser by url    
    elif codeID.lower() == 'title'.lower():   _title(codeID, codeValue) # title() - title of current tagui rpa tab/browser session        
    elif codeID.lower() == 'count'.lower():     _count(codeID, codeValue)   # right click
    elif codeID.lower() == 'select'.lower():    _select(codeValue)  # Selects a dropdown option in a web input. select:dropdown,option
    elif codeID.lower() == 'type'.lower():    _type(codeValue)  # type:identifier,value
    elif codeID.lower() == 'snap'.lower():      _snap(codeValue)    # snap:page,saveFile   Snap entire web page
    elif codeID.lower() == 'telegram'.lower():  _telegram(codeValue)

    elif codeID.lower() == 'download'.lower():    _download(codeValue)          # download - selector, file

    elif codeID.lower() == 'email'.lower():  _email(codeValue, df)
    elif codeID.lower() == 'waitEmailComplete'.lower():  _waitEmailComplete(codeValue, df)
    elif codeID.lower() == 'click'.lower():  _click(codeValue)
    elif codeID.lower() in ['### StartTask'.lower(), '### EndTask'.lower(), '### StartFlow'.lower(), '### EndFlow'.lower()]:  pass
    else:                       _click(code)        # normal left click
    return [], [], []

# used by tagui to redirect console output to a variable
def consoleOutput(runStatement):
    from io import StringIO
    import sys
    old_stdout = sys.stdout # reference current console std out
    buffer = StringIO()     
    sys.stdout = buffer     # redirect std out to buffer
    exec(runStatement)
    buffer_output = buffer.getvalue()   # assign buffer to variable
    sys.stdout = old_stdout   # reset std out
    #sys.stdout = sys.__stdout__
    return buffer_output

from io import StringIO
import sys
def redirectConsole():
    old_stdout = sys.stdout # reference current console std out
    buffer = StringIO()     
    sys.stdout = buffer     # redirect std out to buffer
    return buffer, old_stdout

def resetConsole(buffer, old_stdout):
    buffer_output = buffer.getvalue()   # assign buffer to variable
    sys.stdout = old_stdout   # reset std out
    #sys.stdout = sys.__stdout__
    return buffer_output
    
def _iterate(codeValue, df, objVar):
    '''  Defines iteration steps and appends to runcode list to handle loops over object list or tables.  Syntax: iterate: obj or table list, codelist to run '''
    logger = get_run_logger()
    #logger.info(f">>>>ITERATION ==================== {codeValue}")
    objVar = codeValue.split(',',1)[0].strip()
    sub_code = codeValue.split(',',1)[1].strip()
    worksheetTable = False
    if objVar.lower()[:4]=='tbl@':  # special table object with headers
        withHeader = 1
        objVar = objVar[4:]
    elif objVar in df[(df.Type == 'table')]['Object'].dropna().values.tolist():  # special table object with headers
        withHeader = 1
        #objVar = objVar[4:]
    else:
        # check if its a worksheet name
        result, objTableSet = _isWorkSheetName(objVar, excelfile=STARTFILE)
        if result == True:
            # is a table from a worksheet with name = objVar
            withHeader = 0 #1  Does not support 2 table headers.  Apply a rename instead to align names to function param
            worksheetTable = True
        else:
            # not a table - just normal list.  No header.
            withHeader = 0
    #print('check iterate', objVar, sub_code, withHeader)
    if objVar.isdigit():   # is integer - just standard loop
        totalCount = int(objVar)
        codeBeforeTemplateUpdate = variables['codeBeforeTemplateUpdate']
        sub_code = codeBeforeTemplateUpdate.split(',')[1].strip()
        if len(codeBeforeTemplateUpdate.split(','))==3:
            startCount = codeBeforeTemplateUpdate.split(',')[2].strip()
            startCount = int(startCount) if startCount.isdigit() else 0 
        else:
            startCount = 0
        #print('check .... ' , startCount, totalCount)
        n = 1
        additionalCodeList = []
        additionalDFlist = []
        additionalobjVarList = []
        for i in range(startCount, startCount + totalCount):
            #variables['loopCount'] = i
            #print('ITERATION ---------------------------------------- ', 'LOOP', n)                
            logger.info(f"'ITERATION ---------------------------------------- LOOP {n}")
            #runCode(df, sub_code)
            increment = 'set:loopCount=' + str(i)
            additionalCodeList = additionalCodeList + [increment] + [sub_code]
            additionalDFlist = additionalDFlist + [df, df]
            additionalobjVarList = additionalobjVarList + [objVar, objVar]
            #return [increment, sub_code], [df, df], [objVar, objVar]
            n += 1
        return additionalCodeList, additionalDFlist, additionalobjVarList 
    else:
        #runIterate(df, objVar, sub_code, withHeader)

        # run a sub_code iteratively over a objVar e.g. a URL list like iterate: URL_pages , codeList:<URL_runprocess:key>
        #def runIterate(df, objVar, sub_code, withHeader=0):
 
        if worksheetTable:
            objVarList = objTableSet.iloc[withHeader:, 0].values.tolist() # row and column           
            #logger.info(f"      objTableSet: {objTableSet.iloc[withHeader:3, 0].head(5)}")
            #logger.info(f"      objTableSet: {objTableSet.iloc[withHeader:, 0].values.tolist()}")
        else:
            objVarList = dfObjList(df, objVar, withHeader)
        logger.debug(f"{log_space}LOOP over {codeValue}.  {objVar}:{objVarList}")
        #print(f"      Iteration list: {objVar}, {objVarList}")
        #logg('******** runIterate objVar:', objVar = objVar) # e.g. URL_Dclick_Pages
        #logg('******** runIterate sub_code:', sub_code = sub_code) # e.g. openPages
        #logg('******** runIterate objVarList:', objVarList = objVarList) # e.g. ['wBags', 'wRTW', 'wShoes']
        #logg('******** runIterate codeList:', codeList = codeList) # e.g. codelist of openPages i.e.['print ...' , 'urls...']
        i = 0
        rtn_code = []
        rtn_df = []
        rtn_obj = []
        for x in objVarList:
            #constants['iterationCount'] = i
            # set: iterationCount=i
            rtn_code = rtn_code + [f"iterationCount:{i},{objVarList[i]}"]  # inserts a interationCount step to increase interatation counter
            rtn_df = rtn_df + [df]
            rtn_obj = rtn_obj+ [objVarList[i]]

            #logg('runIterate objVarList[i],', i = i , objVarList = objVarList[i], constantsIterationCount = constants['iterationCount'], level = 'info')
            #print('ITERATION **************************************** ', 'COUNT', i+1, objVarList[i]) #, ', STEP', sub_code)
            #runCode(df, sub_code, objVarList[i])
            rtn_code = rtn_code + [sub_code]
            rtn_df = rtn_df + [df]
            rtn_obj = rtn_obj+ [objVarList[i]]
            i = i + 1
        #logger.info('_iterate complete iteration step preparations.  Next run steps.')
        return rtn_code, rtn_df, rtn_obj



#@task
def _if(codeValue, df, objVar):
    logger = get_run_logger()    
    # codeValue = condition : if true block, if false block or if empty then pass
    condition = codeValue.split(':',1)[0].strip()
    codeBlock = codeValue.split(':',1)[1].strip()
    import re
    match = re.findall( r"'''(.*?)'''", codeBlock)     # Output: ['cats', 'dogs']
    # if codeBlock is encapsulated with ''' ''' then the internal is a statement
    #print(match)
    i = 0
    for item in match:
        codeBlock = codeBlock.replace(f"'''{item}'''", 'match'+str(i))
        i = i+1
    #print(codeBlock)
    args = codeBlock.split('ELSE')
    #args
    if len(args)>0:
        if 'match0' in codeBlock:
            codeBlock1 = args[0].replace('match0', match[0]).strip()
        else:
            codeBlock1 = args[0].strip()
        print(codeBlock1)        
    if len(args)>1:
        if 'match1' in codeBlock:
            codeBlock2 = args[1].replace('match1', match[1]).strip()
        else:
            codeBlock2 = args[1].strip()
        print(codeBlock2)
    else:
        codeBlock2 = 'pass'

    if objVar == None or objVar =="": objVar = " "
    #logg('if condition:', condition = condition, codeBlock = codeBlock)
    logger.debug(f'{log_space}IF {str(eval(condition)).upper()} THEN {codeBlock1} ELSE {codeBlock2}')
    
    if eval(condition):
        #logg('condition is true -----')
        #runCode(df, codeBlock)
        return [codeBlock1], [df], [objVar]
        #codeList = dfObjList(df, codeBlock)
        #runCodelist(df, codeList)
    else:
        if codeBlock2 != 'pass':
            return [codeBlock2], [df], [objVar]
        else:
            return [], [], []

#@task
def _isCodeListTask(df, code, objVar):
    # parameterObjs(df)
    sub_code = dfObjList(df, code)
    #logg('Run Code Block - user defined objects in sheet - ParameterObjs: ', code = code, sub_code = sub_code, level = 'info')
    #runCodelist.with_options(name=code)(df, sub_code)
    #runCodelist(CodeObject(df), sub_code)
    sub_code = ['### StartTask:'+code] + sub_code + ['### EndTask']
    n = len(sub_code)
    logger = get_run_logger()
    logger.debug(f"{log_space}Steps:{sub_code}")
    return sub_code, [df] * n, [objVar] * n

#@flow
def _isCodeListFlow(df, code, objVar):
    # parameterObjs(df)
    sub_code = dfObjList(df, code)
    #logg('Run Code Block - user defined objects in sheet - ParameterObjs: ', code = code, sub_code = sub_code, level = 'info')
    #runCodelist.with_options(name=code)(df, sub_code)
    #runCodelist(CodeObject(df), sub_code)
    sub_code = ['### StartFlow:'+code] + sub_code + ['### EndFlow']
    n = len(sub_code)
    logger = get_run_logger()
    logger.debug(f"{log_space}Steps:{sub_code}")
    return sub_code, [df] * n, [objVar] * n

def _isCodeList(df, code, objVar):
    # parameterObjs(df)
    sub_code = dfObjList(df, code)
    #logg('Run Code Block - user defined objects in sheet - ParameterObjs: ', code = code, sub_code = sub_code, level = 'info')
    #runCodelist.with_options(name=code)(df, sub_code)
    #runCodelist(CodeObject(df), sub_code)
    n = len(sub_code)
    logger = get_run_logger()
    logger.debug(f"{log_space}Steps:{sub_code}")
    return sub_code, [df] * n, [objVar] * n

def _runModule(codeValue, df, objVar):
    logger = get_run_logger()
    sheet = codeValue.split(',')[0]
    num_arg = len(codeValue.split(','))
    if num_arg==1: 
        codeBlock = 'main'
        excelfile = ''
    elif num_arg==2: 
        codeBlock = codeValue.split(',')[1]
        excelfile = ''
    elif num_arg==3:
        codeBlock = codeValue.split(',')[1]
        excelfile = codeValue.split(',')[2]

    #excelfile = '' if len(codeValue.split(','))==1 else codeValue.split(',',1)[1]
    #logg('runModule: ', excelfile = excelfile, sheet = sheet, isExcelFileStringValue = str(excelfile==''))
    from config import SCRIPTS_DIR
    # check if only name given without path
    #if Path('D:\optimus\scripts\serviceNow.xlsm').__str__() == Path('D:\optimus\scripts\serviceNow.xlsm').name
    if excelfile=='': excelfile = STARTFILE
    from pathlib import Path
    if Path(excelfile).__str__() == Path(excelfile).name:  # only name given, add script path
        excelfile = Path(SCRIPTS_DIR, excelfile).resolve().__str__()
    else:
        excelfile = Path(excelfile).resolve().__str__()

    if not Path(excelfile).exists():
        logger.debug(f".... runModule file not exists .... sheet:{sheet} excelfile:{excelfile} df:{df.shape}")
        return [],[],[]

    new_df = readExcelConfig(sheet, excelfile)
    #logger.info(f".... runModule .... new_df sheet:{sheet} excelfile:{excelfile} df:{df.shape}")

    import pandas as pd
    from utility_excel import cacheScripts
    from config import PROGRAM_DIR
    #dfmain, msgStr = cacheScripts(script=Path(startfile).name,df=pd.DataFrame(),program_dir=program_dir, startsheet=startsheet, refresh=bool(update), msgStr='')
    new_df, msgStr = cacheScripts(script='OptimusLib.xlsm',df=new_df,program_dir=PROGRAM_DIR, startsheet='main', refresh=False, msgStr='')
    new_df, msgStr = cacheScripts(script='OptimusLibPublic.xlsm',df=new_df,program_dir=PROGRAM_DIR, startsheet='main', refresh=False, msgStr=msgStr)
    msgStr = f"{log_space}Start file:{excelfile}, Start sheet:{sheet}, Start file name:{Path(excelfile).name}\n{log_space}Scripts files loaded:\n" + msgStr
    logger.debug(f"{msgStr}")

    run_code = dfObjList(new_df, codeBlock)               # run the main code block
    #logger.info(f".... runModule .... run_code:{run_code}")

    #runCodelist.with_options(name=sheet)(df, run_code)
    #runCodelist(CodeObject(df), run_code)
    n = len(run_code)
    return run_code, [new_df] * n, [objVar] * n

#@task(name='sub task', description='sub task description')
def _runTask(codeValue, df, objVar):
    #logger = get_run_logger()
    sheet = codeValue.split(',')[0]
    num_arg = len(codeValue.split(','))
    if num_arg==1: 
        codeBlock = 'main'
        excelfile = ''
    elif num_arg==2: 
        codeBlock = codeValue.split(',')[1]
        excelfile = ''
    elif num_arg==3:
        codeBlock = codeValue.split(',')[1]
        excelfile = codeValue.split(',')[2]

    #excelfile = '' if len(codeValue.split(','))==1 else codeValue.split(',',1)[1]
    #logg('runModule: ', excelfile = excelfile, sheet = sheet, isExcelFileStringValue = str(excelfile==''))
    if excelfile=='': excelfile = STARTFILE
    #logger.info(f".... runModule .... sheet:{sheet} excelfile:{excelfile} df:{df.shape}")

    new_df = readExcelConfig(sheet, excelfile)
    #logger.info(f".... runModule .... new_df sheet:{sheet} excelfile:{excelfile} df:{df.shape}")

    run_code = dfObjList(new_df, codeBlock)               # run the main code block
    #logger.info(f".... runModule .... run_code:{run_code}")

    #runCodelist.with_options(name=sheet)(df, run_code)
    #runCodelist(CodeObject(df), run_code)
    n = len(run_code)
    return run_code, [new_df] * n, [objVar] * n

#@flow(name='sub task', description='sub task description')
def _runFlow(codeValue, df, objVar):
    #logger = get_run_logger()
    sheet = codeValue.split(',')[0]
    num_arg = len(codeValue.split(','))
    if num_arg==1: 
        codeBlock = 'main'
        excelfile = ''
    elif num_arg==2: 
        codeBlock = codeValue.split(',')[1]
        excelfile = ''
    elif num_arg==3:
        codeBlock = codeValue.split(',')[1]
        excelfile = codeValue.split(',')[2]

    #excelfile = '' if len(codeValue.split(','))==1 else codeValue.split(',',1)[1]
    #logg('runModule: ', excelfile = excelfile, sheet = sheet, isExcelFileStringValue = str(excelfile==''))
    if excelfile=='': excelfile = STARTFILE
    #logger.info(f".... runModule .... sheet:{sheet} excelfile:{excelfile} df:{df.shape}")

    new_df = readExcelConfig(sheet, excelfile)
    #logger.info(f".... runModule .... new_df sheet:{sheet} excelfile:{excelfile} df:{df.shape}")

    run_code = dfObjList(new_df, codeBlock)               # run the main code block
    #logger.info(f".... runModule .... run_code:{run_code}")

    #runCodelist.with_options(name=sheet)(df, run_code)
    #runCodelist(CodeObject(df), run_code)
    n = len(run_code)
    return run_code, [new_df] * n, [objVar] * n

#@task
def _codeList(codeValue, df, objVar):
    codeList = codeValue.split(',')
    constants['lastCodelist'] = codeList
    #runCodelist(CodeObject(df), codeList)
    n = len(codeList)
    return codeList, [df] * n, [objVar] * n

#@task
import time
def _wait(codeValue, df, objVar):
    logger = get_run_logger()
    #print('*********************************************************', RPABROWSER)    
    if RPABROWSER==1:
        if codeValue=="" or codeValue==None:    # no parameters
            time_sec=1000        
            time.sleep(time_sec)
            # r.wait(time_sec)
            return [], [], []
        if codeValue[:1]=='{':    # parameters in JSON format
            if codeValue=="" or codeValue==None:
                codeValue = "{}"
            #print('-------------wait_codevalue',codeValue)
            import json
            json_object = json.loads(codeValue)
            if 'time_sec' in json_object:
                time_sec = json_object["time_sec"]
                tmpDict = json_object.pop('time_sec')
            else:
                time_sec = 15000
        else:     # parameters not in JSON format
            tmpDict = parseArguments('time_sec,identifier,run_code,run_code_until',codeValue)  #items = 'wait:15:ID:codeA'
            time_sec = int(tmpDict['time_sec'])*1000
            if 'identifier' in tmpDict:                 # do while identifier is found - r.exist
                #logg('identifier', identifier = tmpDict['identifier'])
                print('wait identifier', tmpDict['identifier'])
                # identifier is a special object list
                if tmpDict['identifier'] in df[(df.Type == 'list')]['Object'].dropna().values.tolist():
                    tmpDict['identifier'] = dfObjList(df, tmpDict['identifier'])
                    #print('wait identifier list', tmpDict['identifier'])
                #logger.debug(f"{log_space}timeout',{time_sec},'selector',{tmpDict['identifier']}")
                timeoutError = not p.wait(time_sec, selector=tmpDict['identifier'])
            else:
                time.sleep(int(time_sec/1000))      # time sec is in milli sec
                timeoutError = False    #not p.wait(time_sec)

        #p.initialize(**json_object)
        if timeoutError: #**tmpDict): #int(codeValue)):
            if 'run_code' in tmpDict:                                           #run code if time out
                tmpDict['run_code'] = dfObjList(df, tmpDict['run_code'])
                print('wait run code list', tmpDict['run_code'])
                logger.debug(f"{log_space}Scenario list:{tmpDict['identifier']} Action list:{tmpDict['run_code']}")

                #logger.debug(f"      Run code from wait: {tmpDict['run_code']}")

                run_code = tmpDict['run_code'] #dfObjList(df, tmpDict['run_code'])
                #logg('      Run code:', run_code = run_code)

                #runCodelist(CodeObject(df), run_code)
                n = len(run_code)
                #logger.debug(f"{log_space}Action:{tmpDict['run_code']}, {n} steps:{run_code} {[objVar] * n}")

                return run_code, [df] * n, [objVar] * n
            else:
                logger.warning(f"{log_space}Time out from waiting ...")
                return [], [], []
        return [], [], []
    else:
        tmpDict = parseArguments('time_sec,identifier,run_code,run_code_until',codeValue)  #items = 'wait:15:ID:codeA'
        time_sec = int(tmpDict['time_sec'])
        if 'identifier' in tmpDict:                 # do while identifier is found - r.exist
            #logg('identifier', identifier = tmpDict['identifier'])
            print('wait identifier', tmpDict['identifier'])
            # identifier is a special object list
            if tmpDict['identifier'] in df[(df.Type == 'list')]['Object'].dropna().values.tolist():
                tmpDict['identifier'] = dfObjList(df, tmpDict['identifier'])
                if 'run_code' in tmpDict:                                           #run code if time out
                    tmpDict['run_code'] = dfObjList(df, tmpDict['run_code'])
                print('wait identifier list', tmpDict['identifier'])
                print('wait run code list', tmpDict['run_code'])

                logger = get_run_logger()
                logger.debug(f"{log_space}Scenario list:{tmpDict['identifier']} Action list:{tmpDict['run_code']}")

                matchBool, index = waitIdentifierExist(tmpDict['identifier'], time_sec, 1, False)         #waitIdentifierExist(identifier, time_seconds, interval) - returns true or false
                if not matchBool:
                    #logg('      Time out from waiting', level = 'warning')                    #raise CriticalAccessFailure("TXT logon window did not appear")
                    logger.warning(f"Time out from waiting ...")
                    return [], [], []
                else:
                    if 'run_code' in tmpDict:                                           #run code if time out
                        print('      Run code from wait:', tmpDict['run_code'][index])

                        run_code = dfObjList(df, tmpDict['run_code'][index])
                        #logg('      Run code:', run_code = run_code)

                        #runCodelist(CodeObject(df), run_code)
                        n = len(run_code)
                        logger.debug(f"{log_space}Action:{tmpDict['run_code'][index]}, {n} steps:{run_code} {[objVar] * n}")

                        return run_code, [df] * n, [objVar] * n
                        #return run_code, [df], [objVar]
                    else:
                        return [], [], []

            else: # not a special object list
                if not waitIdentifierExist(tmpDict['identifier'], time_sec, 5, False):         #waitIdentifierExist(identifier, time_seconds, interval) - returns true or false
                    #logg('      Time out from waiting', level = 'warning')                    #raise CriticalAccessFailure("TXT logon window did not appear")
                    if 'run_code' in tmpDict:                                           #run code if time out
                        run_code = dfObjList(df, tmpDict['run_code'])
                        if 'run_code_until' in tmpDict:
                            #logg('Time out - run code:', run_code = run_code, run_code_until = tmpDict['run_code_until'], level = 'debug')
                            #runCodelist(CodeObject(df), run_code, '', tmpDict['run_code_until'])
                            n = len(run_code)
                            return run_code, [df] * n, [objVar] * n                        
                        else:                                                           
                            #logg('      Time out - run code:', run_code = run_code, level = 'error')
                            #runCodelist(CodeObject(df), run_code)
                            n = len(run_code)
                            return run_code, [df] * n, [objVar] * n
                else:
                    return [], [], []

        # no identifier - normal wait
        else:
            #logg('wait', time_sec = time_sec)
            time.sleep(time_sec)
            # r.wait(time_sec)
            return [], [], []

#@task
def _waitDisappear(codeValue, df, objVar):
    logger = get_run_logger()
    tmpDict = parseArguments('time_sec,identifier,run_code,run_code_until',codeValue)  #items = 'wait:15:ID:codeA'
    time_sec = int(tmpDict['time_sec'])
    logger.debug(f'{log_space}checking 1...')
    if 'identifier' in tmpDict:                 # do while identifier is found - r.exist
        logger.debug(f"{log_space}identifier = {tmpDict['identifier']}")
        if not waitIdentifierDisappear(tmpDict['identifier'], time_sec, 1, False):         #waitIdentifierExist(identifier, time_seconds, interval) - returns true or false
            logger.warning(f"   Time out from waiting', level = 'warning'")                    #raise CriticalAccessFailure("TXT logon window did not appear")
            if 'run_code' in tmpDict:                                           #run code if time out
                run_code = dfObjList(df, tmpDict['run_code'])
                if 'run_code_until' in tmpDict:
                    logger.debug(f"{log_space}Time out - run code: run_code = {run_code}, run_code_until = {tmpDict['run_code_until']}, level = 'debug'")
                    #runCodelist(CodeObject(df), run_code, '', tmpDict['run_code_until'])
                    n = len(run_code)
                    return run_code, [df] * n, [objVar] * n
                else:                                                           
                    logger.warning(f"'      Time out - run code:', run_code = {run_code}, level = 'warning'")
                    #runCodelist(CodeObject(df), run_code)
                    n = len(run_code)
                    return run_code, [df] * n, [objVar] * n
        else:
            return [], [], []
    else:
        logger.debug(f"{log_space}wait time_sec = {time_sec}")
        time.sleep(time_sec)
        # r.wait(time_sec)
        return [], [], []

def _waitFile(codeValue, df, objVar):
    # https://www.geeksforgeeks.org/create-a-watchdog-in-python-to-look-for-filesystem-changes/
    logger = get_run_logger()
    tmpDict = parseArguments('file_pattern, action, timeout, actionIfTimeout',codeValue)  # waitFile: file_pattern, action, timeout, actionIfTimeout
    timeout = int(tmpDict['timeout'])   # timeout in sec
    logger.debug(f'{log_space}checking 1...')

    return [], [], []



#@task
def _print(codeValue):
    #logg('print:', codeValue = codeValue, level = 'info')
    print(codeValue)

def _log(codeValue):
    #logg('log:', errorMsg = codeValue)
    pass

def _exit():
    exitProg()

def _exitError(codeValue):
    if codeValue.isdigit():
        if int(codeValue) in [1,2,3,4,5,6,7,8,9,10]:                
            exitProgWError(errorCode=int(codeValue))

def _raiseError(codeValue):
    raise ValueError(f"Raise Error: {codeValue}")

def _regexSearch(codeValue):
    #regexSearch:PACIFIC.ASIA.(..........),<strSearch>,lastDataUpdate
    strPattern = codeValue.split(',')[0]
    strSearch = codeValue.split(',')[1]
    variable_name = codeValue.split(',')[2]
    variables[variable_name] = regexSearch(strPattern, strSearch)
    #logg('regexSearch', strPattern = strPattern, strSearch = strSearch, variable_name = variable_name, result = variables[variable_name])


def _createPDF(codeValue, df):
    logger = get_run_logger()
    tmpDict = parseArguments('imagelist,outputPath,saveFileName',codeValue)
    if 'imagelist' in tmpDict:
        #logg('imagelist', content = tmpDict['imagelist'].strip())
        imagelist = tmpDict['imagelist'].strip()
        imagelist = dfObjList(df, imagelist)

        #logg('imagelist', imagelist = imagelist)
        #if imagelist == '': imagelist = ''
        if imagelist == None:
            #imagelist = ''
            logger.error('Error - no imagelist ...')
        else:
            if 'outputPath' in tmpDict:
                #logg('outputPath', content = tmpDict['outputPath'])  # D:\\iCristal\\
                outputPath = tmpDict['outputPath'].strip()
                if outputPath == '': outputPath = './' #'.\\' #'D:\\iCristal\\'
            if 'saveFileName' in tmpDict:
                #logg('saveFileName', content = tmpDict['saveFileName'])  # 'D:/iCristal/Output/APAC_Daily_Sales/'
                saveFileName = tmpDict['saveFileName'].strip()
                if saveFileName == '': saveFileName = './savefile.pdf' #'./Output/APAC_Daily_Sales/' #'D:/iCristal/Output/APAC_Daily_Sales/'
            createPDFfromImages(imagelist, outputPath, saveFileName)


def _addContentPDF(codeValue, df):
    tmpDict = parseArguments('pageTitles,sourcePDF,targetPDF',codeValue)
    if 'pageTitles' in tmpDict:
        #logg('pageTitles', content = tmpDict['pageTitles'].strip())
        pageTitles = tmpDict['pageTitles'].strip()
        pageTitlesList = dfObjList(df, pageTitles)
        #logg('pageTitlesList', pageTitlesList = pageTitlesList)
        if not len(pageTitlesList) : return # error - no pageTitles provided
    if 'sourcePDF' in tmpDict:
        #logg('sourcePDF', content = tmpDict['sourcePDF'])  # D:\\iCristal\\
        sourcePDF = tmpDict['sourcePDF'].strip()
        if sourcePDF == '': return # error - no pdf_path defined
    if 'targetPDF' in tmpDict:
        #logg('targetPDF', content = tmpDict['targetPDF'])  # 'D:/iCristal/Output/APAC_Daily_Sales/'
        targetPDF = tmpDict['targetPDF'].strip()
        file_extension = '_' + yesterdayYYYYMMDD + '.pdf'
        if targetPDF == '': targetPDF = sourcePDF + file_extension #'./Output/APAC_Daily_Sales/' #'D:/iCristal/Output/APAC_Daily_Sales/'
    addContentPDF(pageTitlesList, sourcePDF, targetPDF)

def _cropImage(codeValue, df):
    tmpDict = parseArguments('files, savefiles, left, top, right, bottom, boolPercentage',codeValue)
    # cropImage - num or pct - percentage
    tmpDict['files'] = [updateConstants(df, s) for s in dfObjList(df, tmpDict['files'])]
    tmpDict['savefiles'] = [updateConstants(df, s) for s in dfObjList(df, tmpDict['savefiles'])]        
    #print(tmpDict['files'], tmpDict['savefiles'], tmpDict['left'], tmpDict['top'], tmpDict['right'], tmpDict['bottom'], tmpDict['boolPercentage'])
    cropImage(tmpDict['files'], tmpDict['savefiles'], tmpDict['left'], tmpDict['top'], tmpDict['right'], tmpDict['bottom'], tmpDict['boolPercentage'])

def _runExcelMacro(codeValue):
    logger = get_run_logger()
    excel = codeValue.split(',')[0].strip()
    macro = codeValue.split(',')[1].strip()
    import win32com.client
    xl = win32com.client.Dispatch("Excel.Application")          #instantiate excel app
    #xl.Visible = True is not necessary, used just for convenience'

    workBookName = Path(excel).name
    excel = Path(excel).absolute()

    if excel == '': return
    excelQuit = True
    wbClose = True
    if xl.Workbooks.Count > 0:  # excel open with workbooks - do not quit excel
        #for i in office.Workbooks: print('list of open workbooks',i.Name)
        excelQuit = False
        # target workbook is open - do not close the workbook
        if any(i.Name == workBookName for i in xl.Workbooks): 
            #print('do not close workbook')
            wbClose = False

    xl.Visible = False
    logger.debug(f"{log_space}excel {excel.__str__()} workBookName {workBookName.__str__()} macro {macro}")
    wb = xl.Workbooks.Open(excel)
    xl.Application.Run(macro)
    wb.Save()
    if wbClose: wb.Close(SaveChanges=False)
    #xl.Visible = True
    if excelQuit: xl.Application.Quit()


def _runPowerShellScript(codeValue):
    script = codeValue #.split(',')[0].strip()
    #macro = codeValue.split(',')[1].strip()
    import subprocess, sys
    #p = subprocess.Popen(['powershell.exe', '.\AddOn\cleanProcesses.ps1 -msg " Testing 123 v5 " '], stdout=sys.stdout)
    p = subprocess.Popen(['powershell.exe', 
        script], 
        stdout=sys.stdout)
    p.communicate()
    #logg('runPowerShellScript:', returnCode = p.returncode, stderr = p.stderr, stdout = p.stdout)

def _runBatchScript(codeValue):
    script = codeValue.split(',')[0].strip()
    if len(codeValue.split(','))>1:
        workingDir = codeValue.split(',')[1].strip()
    else:
        workingDir = CWD_DIR
    #https://riptutorial.com/python/example/5714/more-flexibility-with-popen
    from subprocess import Popen
    #p = Popen("test.bat", cwd=r".\AddOn")
    p = Popen(script, cwd=workingDir)
    stdout, stderr = p.communicate()
    #logg('runBatchScript:', returnCode = p.returncode, stderr = p.stderr, stdout = p.stdout)
    #print(p.returncode)
    #print("An error occured: %s", p.stderr)
    #print("Output %s", p.stdout)

def _triggerRPAScript(codeValue):
    import yaml
    # write python obj to yaml file
    def write_yaml(py_obj,filename):
        with open(f'{filename}', 'w',) as f :
            yaml.dump(py_obj,f,sort_keys=False) 
        #print('Written to file successfully')
        return True

    # read yaml to python obj/dictionary
    def read_yaml(filename):
        with open(f'{filename}','r') as f:
            output = yaml.safe_load(f)
        print(output)
        return output

    # trigger RPA event by generating token in pending
    def triggerRPA(file='', token={}, memoryPath=''):
        from pathlib import Path, PureWindowsPath
        #print(f"background:{background}")
        state="pending"
        from config import STARTFILE, STARTSHEET, STARTCODE, PROGRAM_DIR, UPDATE, BACKGROUND, RETRIES, MEMORYPATH
        if memoryPath=='': memoryPath=MEMORYPATH
        #write_yaml_to_file(data, 'output.txt')
        if token == {}:
            #token = {}
            token['update']=UPDATE
            token['startfile']=STARTFILE
            token['startsheet']=STARTSHEET
            token['startcode']=STARTCODE
            token['background']=BACKGROUND
            token['program_dir']=PROGRAM_DIR
            print('Token',token)
        print('LAUNCH RPA SCRIPT:', Path(file).stem, write_yaml(token, rf"{memoryPath}\{state}\{Path(file).stem}.txt"))

        result = read_yaml(rf"{memoryPath}\{state}\{Path(file).stem}.txt")
        print(result)
        return True

    # ----------------------------------------
    #script = 'test -sc not ready -b okay -u happy'  # triggerRPAScript:test -u 1 -b 0 -r 3 -sh main -sc main
    script = codeValue.split(',')[0].strip()  # triggerRPAScript:test -u 1 -b 0 -r 3 -sh main -sc main    
    file = codeValue.split(' ')[0].strip()
    #default token values
    from config import STARTFILE, STARTSHEET, STARTCODE, PROGRAM_DIR, UPDATE, BACKGROUND, RETRIES, MEMORYPATH    
    token = {}
    token['update']=UPDATE
    token['startfile']=STARTFILE
    token['startsheet']=STARTSHEET
    token['startcode']=STARTCODE
    token['background']=BACKGROUND
    token['program_dir']=PROGRAM_DIR

    print('Token',token)    
    strSearch = script+' '
    strSearch = strSearch.replace("-u ", "--update ")
    strSearch = strSearch.replace("-b ", "--background ")
    strSearch = strSearch.replace("-r ", "--retries ")

    import re
    # Find all occurrences of single whitespace characters
    #result = re.findall(r'-+([\D]+?)\s+(.*?)\s', string)
    strPattern=r'\s-+([\D]+?)\s+(.*?)\s'
    flags = re.findall(strPattern, strSearch)
    #flags = regexSearch(strPattern=r'-+([\D]+?)\s+(.*?)\s', strSearch=strSearch)    
    print(flags)
    # Python code to merge dict using update() method
    def Merge(dict1, dict2):
        print('merge',dict1, dict2)
        print('MERGED',dict2.update(dict1))
        return(dict2.update(dict1))

    # Python code to convert into dictionary
    def Convert(tup, di):
        for a, b in tup:
            #di.setdefault(a, []).append(b)
            di.setdefault(a, b)
        return di 

    dictionary = {}
    flags = Convert(flags, dictionary)
    print('flags',flags, token,type(flags), type(token) )
    Merge(flags, token)
    print('new token',token)

    print('Triggered', triggerRPA(file=file, token=token))
    #if len(codeValue.split(','))>1:
    #    workingDir = codeValue.split(',')[1].strip()
    #else:
    #    workingDir = CWD_DIR
    #https://riptutorial.com/python/example/5714/more-flexibility-with-popen
    #from subprocess import Popen
    #p = Popen("test.bat", cwd=r".\AddOn")
    #p = Popen(script, cwd=workingDir)
    #stdout, stderr = p.communicate()

def _openProgram(codeValue):
    script = codeValue.split(',')[0].strip()
    if len(codeValue.split(','))>1:
        workingDir = codeValue.split(',')[1].strip()
    else:
        workingDir = CWD_DIR

    # To force the new window to the top
    #from auto_helper_lib import Window
    #selectedWindows = Window()  # instantiate windows object with snapshot of existing windows

    from pywinauto import Application    
    Application().start(script)

    #selectedWindows.getNew()    # get newly opened windows compared to previous snapshot
    #variables['lastWindow']=selectedWindows
    #variables['lastWindow'].focus() # focus the newly opend window with name of    name='google chrome'

def _keyboard_pwa(codeValue:str):
    # from pywinauto.SendKeysCtypes import SendKeys # old for pywinauto==0.5.x
    from pywinauto.keyboard import send_keys
    send_keys(codeValue.replace(" ","{SPACE}"))            
    #send_keys("{VK_SHIFT down}"
    #        "pywinauto"
    #        "{VK_SHIFT up}") # to type PYWINAUTO    

def _focusWindow(codeValue:str):
    from auto_helper_lib import Window
    selectedWins= Window()  # instantiate windows object with snapshot of existing windows
    #selectedWindows.get(name=codeValue)
    selectedWins.focus(name=codeValue) # focus the newly opend window with name of    name='google chrome'     #     

def _focusWindowTemp(codeValue):
    script = codeValue.split(',')[0].strip()
    if len(codeValue.split(','))>1:
        workingDir = codeValue.split(',')[1].strip()
    else:
        workingDir = CWD_DIR
    from pywinauto import Application
    app = Application(backend="uia").connect(path=script)
    app.top_window().set_focus()

def _runJupyterNb(codeValue):
    logger = get_run_logger()
    nb_file = codeValue.split(',', 1)[0].strip()
    jsonString = codeValue.split(',', 1)[1].strip()

    logger.debug(f"{log_space}Run Jupyter Notebook = {nb_file}, Parameters = {jsonString}")

    from auto_initialize import checkWorkDirectory
    from pathlib import Path, PureWindowsPath

    CWD_DIR = checkWorkDirectory('.')
    #logger.info(f"curDIR = {CWD_DIR}")

    #print(jsonString)
    import papermill as pm

    #file = r'C:\Users\roh\Downloads\d5c7a4f7-b9a7-4d1e-904e-ce7349e0f27c.xlsx'
    #country = 'All'

    import json

    #jsonString = '{"file":"C:\\Users\\roh\\Downloads\\d5c7a4f7-b9a7-4d1e-904e-ce7349e0f27c.xlsx", "country": "All"}'
    #jsonString = '{"file": "C:/Users/roh/Downloads/d5c7a4f7-b9a7-4d1e-904e-ce7349e0f27c.xlsx", "country": "All"}'

    paramDict = json.loads(jsonString)
    #logger.info(f"parameter dictionary = {paramDict}")    
    #print(paramDict['file'])
    #print(paramDict['country'])

    #res = pm.execute_notebook(
    #    'C:\\Users\\roh\Documents\\python\\Optimus\\autobot\\DClickReport.ipynb',
    #    'C:\\Users\\roh\\Documents\\python\\Optimus\\autobot\\DClickReport_output.ipynb',
    #    parameters = paramDict
    #)

    from datetime import datetime
    currentDateAndTime = datetime.now()
    #print("The current date and time is", currentDateAndTime)
    # Output: The current date and time is 2022-03-19 10:05:39.482383
    #currentTime = currentDateAndTime.strftime("%H:%M:%S")
    #print("The current time is", currentTime)
    # The current time is 10:06:55

    res = pm.execute_notebook(
        nb_file, nb_file.replace('.ipynb', '_output_'+ currentDateAndTime.strftime("%Y%m%d_%H%M%S") +'.ipynb'),
        parameters = paramDict
    )



def _mergeFiles(codeValue, df):
    # e.g. mergeFiles: mergeFilelist, first, deDuplicationColumnList, ./merged.csv,  latin-1 .... Handles csv or excel
    tmpDict = parseArguments('fileList, keep, uniqueColumnsList, fileName, encoding',codeValue)
    #print(tmpDict)
    if tmpDict == {}: return

    if tmpDict['fileList'] in df[(df.Type == 'list')]['Object'].dropna().values.tolist():           #run Block of Code
        tmpDict['fileList'] = dfObjList(df, tmpDict['fileList'])
        tmpDict['fileList'] = [updateConstants(df, s) for s in tmpDict['fileList']]
        tmpDict['fileList'] = [s.strip() for s in tmpDict['fileList']]        # using list comprehension to trim spaces of every element in list

    if tmpDict['uniqueColumnsList'] in df[(df.Type == 'list')]['Object'].dropna().values.tolist():           #run Block of Code
        tmpDict['uniqueColumnsList'] = dfObjList(df, tmpDict['uniqueColumnsList'])
        tmpDict['uniqueColumnsList'] = [updateConstants(df, s) for s in tmpDict['uniqueColumnsList']]

    #logg('tmpDict', tmpDict = tmpDict, level = 'info')
    #logg('fileList', fileList = tmpDict['fileList'], level = 'info')

    df_merged = pd.concat(map(lambda x: pd.read_csv(x, encoding=tmpDict['encoding']) if x.upper().endswith('.CSV') else pd.read_excel(x), tmpDict['fileList']), ignore_index=True)        
    df_merged = df_merged.drop_duplicates(tmpDict['uniqueColumnsList'], keep=tmpDict['keep'])
    saveFile = tmpDict['fileName']
    df_merged.to_csv(saveFile, index=False) if saveFile.upper().endswith('CSV') else df_merged.to_excel(tmpDict['fileName'], index=False)
    print('      ', 'Merge Files:', tmpDict['fileList'], ' Save to:', saveFile)


def _dropNRowsExcel(codeValue):
    tmpDict = parseArguments('Nrows, variableDataFrameName',codeValue)
    # Drop first N rows
    # by selecting all rows from 4th row onwards
    N = variables[tmpDict['Nrows']]
    variables[tmpDict['variableDataFrameName']] = variables[tmpDict['variableDataFrameName']].iloc[N: , :]

def _DFpromoteHeader(codeValue):
    tmpDict = parseArguments('variableDataFrameName',codeValue)
    new_header = variables[tmpDict['variableDataFrameName']].iloc[0] #grab the first row for the header
    variables[tmpDict['variableDataFrameName']] = variables[tmpDict['variableDataFrameName']][1:] #take the data less the header row
    variables[tmpDict['variableDataFrameName']].columns = new_header #set the header row as the df header

def _DFsaveToExcel(codeValue):
    tmpDict = parseArguments('filename, variableDataFrameName, mode, sheet',codeValue)
    filename = tmpDict['filename']
    dataFrameName = tmpDict['variableDataFrameName']
    if 'sheet' in tmpDict: 
        _sheet = tmpDict['sheet']
    else:
        _sheet = 'Sheet1'
    _ifsheetexist = None
    if 'mode' in tmpDict:
        _mode = tmpDict['mode'] # write w or append a
        if _mode.lower() == 'a': 
            _ifsheetexist = 'overlay'
            append_df_to_excel(filename, variables[dataFrameName], sheet_name=_sheet,
                        index=False)
    else:
        _mode = 'w'  # write
        #variables[tmpDict['variableDataFrameName']].to_excel(tmpDict['filename'], index=False)
        with pd.ExcelWriter(filename, if_sheet_exists=_ifsheetexist, engine="openpyxl",
                            mode=_mode) as writer:  
            variables[dataFrameName].to_excel(writer, sheet_name= _sheet, index=False)
        #    #variables[dataFrameName].to_excel(writer, index=False)
        print(_ifsheetexist, _sheet, _mode)

from pathlib import Path
from copy import copy
from typing import Union, Optional
import numpy as np
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def copy_excel_cell_range(
        src_ws: openpyxl.worksheet.worksheet.Worksheet,
        min_row: int = None,
        max_row: int = None,
        min_col: int = None,
        max_col: int = None,
        tgt_ws: openpyxl.worksheet.worksheet.Worksheet = None,
        tgt_min_row: int = 1,
        tgt_min_col: int = 1,
        with_style: bool = True
) -> openpyxl.worksheet.worksheet.Worksheet:
    """
    copies all cells from the source worksheet [src_ws] starting from [min_row] row
    and [min_col] column up to [max_row] row and [max_col] column
    to target worksheet [tgt_ws] starting from [tgt_min_row] row
    and [tgt_min_col] column.

    @param src_ws:  source worksheet
    @param min_row: smallest row index in the source worksheet (1-based index)
    @param max_row: largest row index in the source worksheet (1-based index)
    @param min_col: smallest column index in the source worksheet (1-based index)
    @param max_col: largest column index in the source worksheet (1-based index)
    @param tgt_ws:  target worksheet.
                    If None, then the copy will be done to the same (source) worksheet.
    @param tgt_min_row: target row index (1-based index)
    @param tgt_min_col: target column index (1-based index)
    @param with_style:  whether to copy cell style. Default: True

    @return: target worksheet object
    """
    if tgt_ws is None:
        tgt_ws = src_ws

    # https://stackoverflow.com/a/34838233/5741205
    for row in src_ws.iter_rows(min_row=min_row, max_row=max_row,
                                min_col=min_col, max_col=max_col):
        for cell in row:
            tgt_cell = tgt_ws.cell(
                row=cell.row + tgt_min_row - 1,
                column=cell.col_idx + tgt_min_col - 1,
                value=cell.value
            )
            if with_style and cell.has_style:
                # tgt_cell._style = copy(cell._style)
                tgt_cell.font = copy(cell.font)
                tgt_cell.border = copy(cell.border)
                tgt_cell.fill = copy(cell.fill)
                tgt_cell.number_format = copy(cell.number_format)
                tgt_cell.protection = copy(cell.protection)
                tgt_cell.alignment = copy(cell.alignment)
    return tgt_ws

#   append_df_to_excel(filename, variables[dataFrameName], sheet_name=_sheet, index=False)
def append_df_to_excel_old(
    filename: Union[str, Path],
    df: pd.DataFrame,
    sheet_name: str = 'Sheet1',
    startrow: Optional[int] = None,
    max_col_width: int = 30,
    autofilter: bool = False,
    fmt_int: str = "#,##0",
    fmt_float: str = "#,##0.00",
    fmt_date: str = "yyyy-mm-dd",
    fmt_datetime: str = "yyyy-mm-dd hh:mm",
    truncate_sheet: bool = False,
    storage_options: Optional[dict] = None,
    **to_excel_kwargs
) -> None:
    import pandas
    from openpyxl import load_workbook

    book = load_workbook(filename)
    writer = pandas.ExcelWriter(filename, engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    for sheetname in writer.sheets:
        df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)

    writer.save()


def append_df_to_excel(
        filename: Union[str, Path],
        df: pd.DataFrame,
        sheet_name: str = 'Sheet1',
        startrow: Optional[int] = None,
        max_col_width: int = 30,
        autofilter: bool = False,
        fmt_int: str = "#,##0",
        fmt_float: str = "#,##0.00",
        fmt_date: str = "yyyy-mm-dd",
        fmt_datetime: str = "yyyy-mm-dd hh:mm",
        truncate_sheet: bool = False,
        storage_options: Optional[dict] = None,
        **to_excel_kwargs
) -> None:
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param max_col_width: maximum column width in Excel. Default: 40
    @param autofilter: boolean - whether add Excel autofilter or not. Default: False
    @param fmt_int: Excel format for integer numbers
    @param fmt_float: Excel format for float numbers
    @param fmt_date: Excel format for dates
    @param fmt_datetime: Excel format for datetime's
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param storage_options: dict, optional
        Extra options that make sense for a particular storage connection, e.g. host, port,
        username, password, etc., if using a URL that will be parsed by fsspec, e.g.,
        starting “s3://”, “gcs://”.
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('/tmp/test.xlsx', df, autofilter=True,
                           freeze_panes=(1,0))

    >>> append_df_to_excel('/tmp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('/tmp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('/tmp/test.xlsx', df, sheet_name='Sheet2',
                           index=False, startrow=25)

    >>> append_df_to_excel('/tmp/test.xlsx', df, index=False,
                           fmt_datetime="dd.mm.yyyy hh:mm")

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    def set_column_format(ws, column_letter, fmt):
        for cell in ws[column_letter]:
            cell.number_format = fmt
    filename = Path(filename)
    file_exists = filename.is_file()
    # process parameters
    # calculate first column number
    # if the DF will be written using `index=True`, then `first_col = 2`, else `first_col = 1`
    first_col = int(to_excel_kwargs.get("index", True)) + 1
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    # save content of existing sheets
    if file_exists:
        wb = load_workbook(filename)
        sheet_names = wb.sheetnames
        sheet_exists = sheet_name in sheet_names
        sheets = {ws.title: ws for ws in wb.worksheets}

    with pd.ExcelWriter(
        filename.with_suffix(".xlsx"),
        engine="openpyxl",
        mode="a" if file_exists else "w",
        if_sheet_exists="new" if file_exists else None,
        date_format=fmt_date,
        datetime_format=fmt_datetime,
        storage_options=storage_options
    ) as writer:
        if file_exists:
            # try to open an existing workbook
            writer.book = wb
            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row
            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)
            # copy existing sheets
            writer.sheets = sheets
        else:
            # file doesn't exist, we are creating a new one
            startrow = 0

        # write out the DataFrame to an ExcelWriter
        df.to_excel(writer, sheet_name=sheet_name, **to_excel_kwargs)
        worksheet = writer.sheets[sheet_name]

        if autofilter:
            worksheet.auto_filter.ref = worksheet.dimensions

        for xl_col_no, dtyp in enumerate(df.dtypes, first_col):
            col_no = xl_col_no - first_col
            width = max(df.iloc[:, col_no].astype(str).str.len().max(),
                        len(df.columns[col_no]) + 6)
            width = min(max_col_width, width)
            column_letter = get_column_letter(xl_col_no)
            worksheet.column_dimensions[column_letter].width = width
            if np.issubdtype(dtyp, np.integer):
                set_column_format(worksheet, column_letter, fmt_int)
            if np.issubdtype(dtyp, np.floating):
                set_column_format(worksheet, column_letter, fmt_float)

    if file_exists and sheet_exists:
        # move (append) rows from new worksheet to the `sheet_name` worksheet
        wb = load_workbook(filename)
        # retrieve generated worksheet name
        new_sheet_name = set(wb.sheetnames) - set(sheet_names)
        if new_sheet_name:
            new_sheet_name = list(new_sheet_name)[0]
        # copy rows written by `df.to_excel(...)` to
        copy_excel_cell_range(
            src_ws=wb[new_sheet_name],
            tgt_ws=wb[sheet_name],
            tgt_min_row=startrow + 1,
            with_style=True
        )
        # remove new (generated by Pandas) worksheet
        del wb[new_sheet_name]
        wb.save(filename)
        wb.close()

def _DFreadExcel(codeValue):
    tmpDict = parseArguments('filename, variableDataFrameName',codeValue)
    variables[tmpDict['variableDataFrameName']] = pd.read_excel(tmpDict['filename'])

def _DFsort(codeValue):
    tmpDict = parseArguments('sortFieldList, boolAscending, variableDataFrameName',codeValue)

    columnList = tmpDict['sortFieldList']
    boolAscending = tmpDict['boolAscending']
    variables[tmpDict['variableDataFrameName']].sort_values(by=columnList, ascending = boolAscending, inplace = True)

def _DFdropDuplicates(codeValue):
    tmpDict = parseArguments('uniqueColumnList, keep, variableDataFrameName',codeValue)

    uniqueColumnList = tmpDict['uniqueColumnList']
    keep = tmpDict['keep']
    variables[tmpDict['variableDataFrameName']] = variables[tmpDict['variableDataFrameName']].drop_duplicates(uniqueColumnList, keep=keep)

def _DFconcatenate(codeValue):
    tmpDict = parseArguments('dataFrameList, mergedDataFrameName',codeValue)
    dataFrameList = tmpDict['dataFrameList']
    variables[tmpDict['mergedDataFrameName']] = pd.concat(map(lambda x: variables[x], dataFrameList), ignore_index=True)

def _DFcreate(codeValue):
    tmpDict = parseArguments('dataFrameName, columnslist',codeValue)
    dataFrameName = tmpDict['dataFrameName']
    columnslist = tmpDict['columnslist'].strip('\"')
    elementlist = list(map(lambda x: x.strip(), columnslist.split(',')))
    dictionary = variables
    print('1', dataFrameName, elementlist, dictionary)
    dictresult = dict((k, dictionary[k]) for k in elementlist
            if k in dictionary)
    print('2', dictresult)
    if not dataFrameName in variables:
        #n == 0:
        variables[dataFrameName] = pd.DataFrame(dictresult, index=[0])
    else:
        n = variables[dataFrameName].__len__()
        variables[dataFrameName] = pd.concat([variables[dataFrameName], pd.DataFrame(dictresult, index=[n])])
    print('3', variables[dataFrameName])

def _chromeZoom(codeValue):
    chromeZoom(codeValue)

def _copyFile(codeValue):
    source = codeValue.split(',',1)[0]
    dest = codeValue.split(',',1)[1]
    #logg('copyFile:', source = source, dest = dest)
    import os
    import shutil
    destination = shutil.copyfile(source, dest)
    #logg("Path of copied file:", destination = destination)

def _moveFile(codeValue):
    source = codeValue.split(',',1)[0]
    dest = codeValue.split(',',1)[1]
    #logg('moveFile:', source = source, dest = dest)
    import os
    import shutil
    destination = shutil.move(source, dest)
    #logg("Path of moved file:", destination = destination)

def _makeDir(codeValue):
        path = codeValue.split(',',1)[0]
        import pathlib
        p = pathlib.Path(path)
        p.mkdir(parents=True, exist_ok=True)
        print("      ","Directory created", path)

def _removeFile(codeValue):
    filePattern = codeValue
    #logg('removeFile:', filePattern = filePattern)
    import os, glob
    # Getting All Files List
    fileList = glob.glob(filePattern, recursive=True)
    # Remove all files one by one
    for file in fileList:
        try:
            os.remove(file)
        except OSError:
            print("      ","Error while deleting file")
    #logg("Removed all matched files!")
    print("      ","Removed all matched files!")


#@task
def _renameRecentDownloadFile(codeValue):
    # e.g. codeValue = 'renameRecentDownloadFile:D:/iCristal/Output/APAC_Daily_Sales/,D:\\iCristal\\,*.pdf,120'
    tmpDict = parseArguments('saveName,saveName_suffix,sourcePath,targetPath,fileExtension,withinLastSec',codeValue)
    print('renameRecentDownloadFile','saveName,saveName_suffix,sourcePath,targetPath,fileExtension,withinLastSec', tmpDict)
    if 'saveName' in tmpDict:
        #logg('saveName', saveName = tmpDict['saveName'].strip())
        saveName = tmpDict['saveName'].strip()
        if saveName == '': saveName = variables['url']
    if 'saveName_suffix' in tmpDict:
        #logg('saveName_suffix',saveName_suffix = tmpDict['saveName_suffix'].strip())
        saveName_suffix = tmpDict['saveName_suffix'].strip()
        if saveName_suffix == '': saveName_suffix = ''
    if 'sourcePath' in tmpDict:
        #logg('sourcePath',sourcePath = tmpDict['sourcePath'])  # D:\\iCristal\\
        sourcePath = tmpDict['sourcePath'].strip()
        if sourcePath == '': sourcePath = './' #'.\\' #'D:\\iCristal\\'
    if 'targetPath' in tmpDict:
        #logg('targetPath', targetPath = tmpDict['targetPath'])  # 'D:/iCristal/Output/APAC_Daily_Sales/'
        targetPath = tmpDict['targetPath'].strip()
        if targetPath == '': targetPath = './' #'./Output/APAC_Daily_Sales/' #'D:/iCristal/Output/APAC_Daily_Sales/'
    if 'fileExtension' in tmpDict:
        #logg('fileExtension', fileExtension = tmpDict['fileExtension'])
        fileExtension = tmpDict['fileExtension'].strip()
        if fileExtension == '': fileExtension = 'pdf'
    if 'withinLastSec' in tmpDict:
        #logg('withinLastSec', withinLastSec = tmpDict['withinLastSec'])
        withinLastSec = tmpDict['withinLastSec'].strip()
        if withinLastSec == '':
            withinLastSec = 12000
        else:
            withinLastSec = int(withinLastSec)
    saveName = saveName + saveName_suffix
    downloadedFile = GetRecentCreatedFile(sourcePath,'*.' + fileExtension, withinLastSec) # get most recent file of pdf format in last 120 sec in path
    #logg('renameFile parameters', targetPath = targetPath, saveName = saveName, fileExtension = fileExtension, level = 'debug')
    #logg('Recent Download file : none', level = 'error') if downloadedFile is None else #logg('Recent Download file:', downloadFile = downloadedFile, targetPath = targetPath, saveName = saveName, fileExtension = fileExtension, level = 'info')
    if downloadedFile is not None: renameFile(downloadedFile, targetPath + saveName + '.' + fileExtension)

def _touchFile(codeValue):
    # touchFile: file, path optional
    # e.g. codeValue = 'renameRecentDownloadFile:D:/iCristal/Output/APAC_Daily_Sales/,D:\\iCristal\\,*.pdf,120'
    tmpDict = parseArguments('file, path',codeValue)
    print('touchFile','file,path', tmpDict)

    from pathlib import Path
    #_path = "D:\OneDrive-Sync\OneDrive - Christian Dior Couture\Shared Documents - RPA Project-APAC_FIN\Daily Report" # "." # path/to/
    #Path(_path, 'status.txt').touch()
    
    if 'path' in tmpDict:
        _path = tmpDict['path']
    else:
        _path = ""
    Path(_path, tmpDict['file']).touch()

    #filename = codeValue.split(',',1)[0].strip()
    #msg = codeValue.split(',',1)[1].strip()
    #flow_run_name = context.get_run_context().flow_run.dict()['name']  # error if this is not executed in a flow
    #from auto_core_lib import touchFile 
    #r.telegram(int(id),f"{flow_run_name}-{msg}")
    #print(touchFile(filename))

    #if 'saveName' in tmpDict:

    #source = codeValue.split(',',1)[0]
    #dest = codeValue.split(',',1)[1]
    #logg('moveFile:', source = source, dest = dest)
    #import os
    #import shutil
    #destination = shutil.move(source, dest)
    #logg("Path of moved file:", destination = destination)

def _raiseEvent(codeValue):
    # touchFile: file, path optional
    # e.g. codeValue = 'renameRecentDownloadFile:D:/iCristal/Output/APAC_Daily_Sales/,D:\\iCristal\\,*.pdf,120'
    tmpDict = parseArguments('event',codeValue)
    print('event', tmpDict)
    if 'event' in tmpDict:
        from config import MEMORYPATH        
        filename = f"{MEMORYPATH}/event/{tmpDict['event']}.txt" #'/home/tuhingfg/Desktop'        
        #msg = codeValue.split(',',1)[1].strip()
        #flow_run_name = context.get_run_context().flow_run.dict()['name']  # error if this is not executed in a flow
        from auto_core_lib import touchFile 
        #r.telegram(int(id),f"{flow_run_name}-{msg}")
        print('event raised', touchFile(filename))

def _ifEvent(codeValue, df, objVar):
    tmpDict = parseArguments('event, action1, action2',codeValue)
    print('event','action1','action2', tmpDict)
    if 'event' in tmpDict and 'action1' in tmpDict:     
        # Import Path class
        from pathlib import Path
        from config import MEMORYPATH
        # Path
        path = f"{MEMORYPATH}/event/{tmpDict['event']}.txt" #'/home/tuhingfg/Desktop'
        
        # Instantiate the Path class
        obj = Path(path)
        
        # Check if path exists
        run_code = [tmpDict['action1']]
        if 'action2' in tmpDict:
            run_code_else = [tmpDict['action2']]
        else:
            run_code_else = []
        print(f"ifEvent: {tmpDict['event']} : {obj.exists()}, {run_code} ELSE {run_code_else}")
        if obj.exists():
            obj.unlink(missing_ok=False)  #if missing_ok is false (the default), FileNotFoundError is raised if the path does not exist.
            n = len(run_code)
            print('action1', run_code, n)
            return run_code, [df] * n, [objVar] * n
        else:
            n = len(run_code_else)
            print('action2', run_code_else, n)
            return run_code_else, [df] * n, [objVar] * n            
    else:
        return [], [], []        

def _ifTest(codeValue, df, objVar):
    logger = get_run_logger()    
    # codeValue = condition : if true block, if false block or if empty then pass
    condition = codeValue.split(':',1)[0].strip()
    codeBlock = codeValue.split(':',1)[1].strip()
    import re
    match = re.findall( r"'''(.*?)'''", codeBlock)     # Output: ['cats', 'dogs']
    # if codeBlock is encapsulated with ''' ''' then the internal is a statement
    #print(match)
    i = 0
    for item in match:
        codeBlock = codeBlock.replace(f"'''{item}'''", 'match'+str(i))
        i = i+1
    #print(codeBlock)
    args = codeBlock.split('ELSE')
    #args
    if len(args)>0:
        if 'match0' in codeBlock:
            codeBlock1 = args[0].replace('match0', match[0]).strip()
        else:
            codeBlock1 = args[0].strip()
        print(codeBlock1)        
    if len(args)>1:
        if 'match1' in codeBlock:
            codeBlock2 = args[1].replace('match1', match[1]).strip()
        else:
            codeBlock2 = args[1].strip()
        print(codeBlock2)
    else:
        codeBlock2 = 'pass'

    if objVar == None or objVar =="": objVar = " "
    #logg('if condition:', condition = condition, codeBlock = codeBlock)
    logger.debug(f'{log_space}IF {str(eval(condition)).upper()} THEN {codeBlock1} ELSE {codeBlock2}')
    
    if eval(condition):
        #logg('condition is true -----')
        #runCode(df, codeBlock)
        return [codeBlock1], [df], [objVar]
        #codeList = dfObjList(df, codeBlock)
        #runCodelist(df, codeList)
    else:
        if codeBlock2 != 'pass':
            return [codeBlock2], [df], [objVar]
        else:
            return [], [], []


def _runInBackground():
    #script = codeValue.split(',')[0].strip()
    #workingDir = codeValue.split(',')[1].strip()
    runInBackground()

#@task
def _initializeRPA(codeValue: str):
    # visual_automation = False, chrome_browser = True, headless_mode = False, turbo_mode = False
    logger = get_run_logger()

    if RPABROWSER == 1:
        if codeValue=="" or codeValue==None:
            codeValue = "{}"
        #logger.warning(f'{log_space}InitializeRPA codeValue{codeValue}================')
        try:
            import json
            json_object = json.loads(codeValue)
        except Exception as error:
            logger.error('{0}Initialization Error: {1} | {2}'.format(log_space, type(error).__name__, error))
            #logger.debug(log_space+traceback.format_exc())

        var_exists = 'p' in globals() #'p' in locals() or 'p' in globals()
        global p
        if var_exists:
            del p
            p_exists = 'p' in globals() #'p' in locals() or 'p' in globals()            
            logger.error("Deleted object p exist " + p_exists.__str__())
        p = Browser()
        p.initialize(**json_object)
        p_exists = 'p' in globals() #'p' in locals() or 'p' in globals()            
        logger.error("initialized object p exist " + p_exists.__str__())
    else:

        from auto_helper_lib import Window, process_list
        processResult = process_list(name='', minutes=5)
        #selectedWindows = windows_getTitle(name='')
        selectedWindows = Window()  # instantiate windows object with snapshot of existing windows
        #logger.debug(f'{log_space}Windows: {selectedWindows.title}')

        #if not browserDisable:
        #r.init()
        # init(visual_automation = False, chrome_browser = True, headless_mode = False, turbo_mode = False):
        #instantiatedRPA = r.init(visual_automation = True)

        if variables['headless_mode']==True:
            visual_automation = False
            chrome_browser = True
            headless_mode = True
            turbo_mode = False
        else: # variables['headless_mode']==False:
            visual_automation = True #False
            chrome_browser = True
            headless_mode = False
            turbo_mode = False

        jsonString = codeValue.strip()
        #jsonString='{"visual_automation":False, "chrome_browser":True, "headless_mode":True, "turbo_mode":False}' # overwrite setting in command

        import json
        #jsonString = '{"file":"C:\\Users\\roh\\Downloads\\d5c7a4f7-b9a7-4d1e-904e-ce7349e0f27c.xlsx", "country": "All"}'
        #jsonString = '{"file": "C:/Users/roh/Downloads/d5c7a4f7-b9a7-4d1e-904e-ce7349e0f27c.xlsx", "country": "All"}'
        if jsonString == '':
            pass
        else:
            try:
                paramDict = json.loads(jsonString.lower())
                #logger.info(f"parameter dictionary = {paramDict}")    
                #print(paramDict['file'])
                #print(paramDict['country'])
                #print(jsonString, paramDict)
                logger.debug(f"{log_space}RPA Initialize Parameters = {jsonString}, {paramDict}")
                for item in paramDict:
                    if item == "visual_automation": visual_automation=paramDict[item]
                    elif item == "chrome_browser": chrome_browser=paramDict[item]
                    elif item == "headless_mode": headless_mode=paramDict[item]
                    elif item == "turbo_mode": turbo_mode=paramDict[item]
                    #print(item)
                    #print(f"{item}={paramDict[item]}")
                    #exec(f"{item}={paramDict[item]}")  # modify variables
                print(visual_automation, chrome_browser, headless_mode, turbo_mode, type(visual_automation))
            except:
                print("Error json string in _initializeRPA:", jsonString)

        instantiatedRPA = r.init(visual_automation = visual_automation, chrome_browser = chrome_browser, headless_mode = headless_mode, turbo_mode = turbo_mode)
        print(f"r.init(visual_automation = {visual_automation}, chrome_browser = {chrome_browser}, headless_mode = {headless_mode}, turbo_mode = {turbo_mode})")
        #logg('Initialize RPA', result = instantiatedRPA, level = 'info')
        logger.debug(f"{log_space}Initialize RPA = {instantiatedRPA}")

        if instantiatedRPA:
            selectedWindows.getNew()    # get newly opened windows compared to previous snapshot
            title=r.title()
            if title=='': title='about:blank'
            selectedWindows.focus(name=title) #'google chrome') # focus the newly opend window with name of

#@task
def _closeRPA():
    logger = get_run_logger()
    #if not browserDisable:
    if RPABROWSER==0:
        instantiatedRPA = r.close()    
    else:
        p.close_browser()
        instantiatedRPA = True
    logger.debug(f"{log_space}Close RPA = {instantiatedRPA}")
    #logg('Close RPA ', result = instantiatedRPA, level = 'info')

#@task
def _url(codeValue, df):
    logger = get_run_logger()

    tmpDict = parseArguments('key, authentication',codeValue)
    #print(tmpDict)
    if tmpDict == {}: return

    if 'key' in tmpDict:
        key = tmpDict['key'].strip()   #codeValue.strip()
    #print('check key', key)

    if isinstance(key, str) and key[:4]=='http':  # value given is a URL value
        #logger.info(f"      Valid URL")
        url_value = key
    elif key[:1]=='@':                            # value given is a URL value with special prefix @
        url_value = key[1:]
    else:                                         # name of list (key value pair) expected
        #print('check df',df)
        url_value = dfKey_value(df, key)
        #print('      ','URL:',key, url_value) #, type(url_value))
        #logger.info(f"      DEBUG url: ', VARIABLE_TYPE = {type(url_value)}, key = {key}, url_value = {url_value}")
        logger.debug(f"{log_space}Open URL: {url_value}")

    import math
    #x = float('nan')
    #math.isnan(x)

    if url_value==None or url_value!=url_value:  #df key value returns empty value
        logger.info(f"      Not a valid key value pair. url_value = {url_value} key = {key}, level = 'warning'")            
    elif not(isinstance(url_value, str)):
        url_value = url_value[0]
        #logg('IsInstanceOfStr', key = key, url_value = url_value)
    if isinstance(url_value, str) and (url_value[:4]=='http' or url_value[:6]=='chrome'):
        #runStatement = \
        #                '''
        #                result = r.url(url_value) # true if run is successful
        #                '''
        #print(consoleOutput(runStatement))
        buffer, old_stdout = redirectConsole()

        if RPABROWSER == 1:
            if 'authentication' in tmpDict:
                p.page_goto(url_value, authentication=1)
            else:
                p.page_goto(url_value)
            result = True
        else:
            result = r.url(url_value) # true if run is successful
            consoleOutput = resetConsole(buffer, old_stdout)
        if not result:
            logger.info(f"      RPA ERROR: {consoleOutput}")
        else:
            #logger.info(f"      Valid URL")
            #pass
            if RPABROWSER == 0:
                from auto_helper_lib import Window
                selectedWindows = Window()  # instantiate windows object with snapshot of existing windows
                #selectedWindows.getNew()    # get newly opened windows compared to previous snapshot
                title=r.title()
                if title=='': title='about:blank'
                print('title of window to focus', title)
                selectedWindows.focus(name=title) #'google chrome') # focus the newly opend window with name of

    else:
        logger.info(f"      ERROR: Not http address, url_value = {url_value}, key = {key}, level = 'warning'")
    constants['url'] = key          # set the URL key/name to a temp varaible with label url

def _urls(codeValue, objVar):
    #logg('urls',codeValue = codeValue, objVar = objVar)
    url = dfKey_value(df, objVar).strip()                
    #logg('url', objVar = objVar, url = url)
    r.url(url)
    constants['url'] = objVar.strip()

#@task
def _read(codeValue):
    key = codeValue.split('=',1)[0]
    value = codeValue.split('=',1)[1]
    #logg('read:', key = key, value = value)
    if RPABROWSER == 1:
        variables[key] = p.read(value)
        print('    READ', variables[key])
    else:
        variables[key] = r.read(value)  # print('variables(', key, ')=', variables[key])

#@task
def _checkVariable(codeValue):
    logger = get_run_logger()
    variableValue = variables[codeValue]
    logger.debug(f"{log_space}checkVariable:', key = {codeValue}, variableValue = {variableValue}")

#@task
def _set(codeValue):
    key = codeValue.split('=',1)[0].strip()
    value = codeValue.split('=',1)[1].strip()
    if key == 'iterationCount':
        # special case iteration count
        constants[key] = int(value)
    else:
        variables[key] = value
        #logg('set: ', key = key, variablesValue = variables[key])        

def _increment(codeValue):
    key = codeValue.split(',',1)[0].strip()
    value = codeValue.split(',',1)[1].strip()
    if key == 'iterationCount':
        # special case iteration count
        constants[key] = int(value) + int(constants[key])
    else:
        variables[key] = int(value) + int(variables[key])
        #logg('set: ', key = key, variablesValue = variables[key])        

#@task
def _iterationCount(codeValue, df, objVar):
    # ['iterationCount:' + str(i) + ',' + objVarList[i]]
    countValue = codeValue.split(',',1)[0].strip()
    #objVarListDesc = codeValue.split(',',1)[1].strip()
    constants['iterationCount'] = int(countValue)  
    logger = get_run_logger()
    logger.info(f">>>>>>>>>>>>>>>>>>>> ITERATION:{int(countValue)+1}, {objVar} <<<<<<<<<<<<<<<<<")

def _loopWhile(codeValue, df, objVar):                # while:condition:do this:repeat every X sec
    logger = get_run_logger()    
    # codeValue = condition : if true block, if false block or if empty then pass
    condition = codeValue.split(':',1)[0].strip()
    codeBlock = codeValue.split(':',1)[1].strip()
    #repeatInSec = codeBlock.split(':',1)[1].strip()    
    import re
    match = re.findall( r"'''(.*?)'''", codeBlock)     # Output: ['cats', 'dogs']
    # if codeBlock is encapsulated with ''' ''' then the internal is a statement
    #print(match)
    i = 0
    for item in match:
        codeBlock = codeBlock.replace(f"'''{item}'''", 'match'+str(i))
        i = i+1
    #print(codeBlock)
    args = codeBlock.split('ELSE')
    #args
    if len(args)>0:
        if 'match0' in codeBlock:
            codeBlock1 = args[0].replace('match0', match[0]).strip()
        else:
            codeBlock1 = args[0].strip()
        print(codeBlock1)        
    if len(args)>1:
        if 'match1' in codeBlock:
            codeBlock2 = args[1].replace('match1', match[1]).strip()
        else:
            codeBlock2 = args[1].strip()
        print(codeBlock2)
    else:
        codeBlock2 = 'pass'

    if objVar == None or objVar =="": objVar = " "
    #logg('if condition:', condition = condition, codeBlock = codeBlock)
    codeToRun = eval(f"['{codeBlock1}']+['wait:5']")
    codeBlock1 = 'codeList:' + codeBlock1 + ',wait:5, if:{condition}:{codeBlock1} ELSE {codeBlock2}' #codeValue #'loopWhile:' + codeValue
    logger.debug(f'{log_space}IF {str(eval(condition)).upper()} THEN {codeToRun} ELSE {codeBlock2} | {codeValue} | {codeBlock1}')
    
    if eval(condition):
        #logg('condition is true -----')
        #runCode(df, codeBlock)
        import time
        time_sec = 10
        #time.sleep(time_sec)
        #return [codeBlock1], [df], [objVar]
        return [codeBlock1], [df], [objVar]
        #codeList = dfObjList(df, codeBlock)
        #runCodelist(df, codeList)
    else:
        if codeBlock2 != 'pass':
            return [codeBlock2], [df], [objVar]
        else:
            return [], [], []


def _urlcontains(codeValue):
    search = codeValue.split('=',1)[0]
    result_name = codeValue.split('=',1)[1]
    variables[result_name] = search in r.url()
    #logg('urlcontains: ', search = search, result_name = variables[result_name])        
    print('      ',codeID,codeValue,variables[result_name])

#@task
def _rclick(codeValue):
    if RPABROWSER == 1:
        p.click(codeValue, button="right")        
    else:
        if codeValue.lower().endswith(('.png', '.jpg', '.jpeg')): codeValue = Path(IMAGE_DIR + '/' + codeValue).absolute().__str__() 
        r.rclick(codeValue)             # print('rclick',prefix[1])
        print('      ','rclick',codeValue)

#@task
def _present(codeID, codeValue):
    variables[codeID] = r.present(codeValue)             # print('rclick',prefix[1])
    print('      ',codeID,variables[codeID])

#@task
def _exist(codeID, codeValue):
    variables[codeID] = r.exist(codeValue)             # print('rclick',prefix[1])
    print('      ',codeID,variables[codeID])

def _count(codeID, codeValue):
    variables[codeID] = r.count(codeValue)             # print('rclick',prefix[1])
    print('      ',codeID,variables[codeID])

#_focus(codeID, codeValue) # focus() - app_to_focus (full name of app) - make application in focus
def _focus(codeID, codeValue):
    variables[codeID] = r.focus(codeValue)             # print('rclick',prefix[1])
    print('      ',codeID,variables[codeID])

def _popup(codeID, codeValue):
    variables[codeID] = r.popup(codeValue)             # print('rclick',prefix[1])
    print('      ',codeID,variables[codeID])

def _title(codeID, codeValue):
    variables[codeID] = r.title()             # print('rclick',prefix[1])
    print('      ',codeID,variables[codeID])        

#@task
def _keyboard(codeValue: str):
    if RPABROWSER == 1:
        def mapKeyCodes(test_str):
            # Replace Different characters in String at Once using regex + lambda
            import re
            
            # initializing string
            #test_str = '[ctrl][alt][shift][esc][esc][esc][enter][shift][enter]'
            
            # printing original String
            print("The original string is : " + str(test_str))
            
            # initializing mapping dictionary
            # key press e.g. [home] [end] [insert] [f1] .. [f15] [shift] [ctrl] [alt] [win] [cmd] [enter] [space] [tab] [esc] [backspace] [delete] [clear]
            # F1 - F12, Digit0- Digit9, KeyA- KeyZ, Backquote, Minus, Equal, Backslash, Backspace, Tab, Delete, Escape, ArrowDown, End, Enter, Home, Insert, PageDown, PageUp, ArrowRight, ArrowUp, etc.
            # Following modification shortcuts are also supported: Shift, Control, Alt, Meta, ShiftLeft.
            '''
            [shift] [ctrl] [alt] [win] [cmd] [enter]
            [space] [tab] [esc] [backspace] [delete] [clear]
            [up] [down] [left] [right] [pageup] [pagedown]
            [home] [end] [insert] [f1] .. [f15]
            [printscreen] [scrolllock] [pause] [capslock] [numlock]
            https://developer.mozilla.org/en-US/docs/Web/API/UI_Events/Keyboard_event_key_values
            https://tagui.readthedocs.io/en/latest/reference.html
            https://playwright.dev/docs/api/class-keyboard
            https://github.com/tebelorg/RPA-Python#core-functions
            https://www.geeksforgeeks.org/python-replace-different-characters-in-string-at-once/
            '''
            map_dict1 = {'\[':'[','\]':']',
                        '\-':'Minus','\+':'Plus','\=':'Equal',
                    }

            map_dict = {'\[':'[','\]':']',
                        'shift':'Shift+','ctrl':'Control+','alt':'Alt+','win':'Meta','cmd':'Meta','enter':'Enter',
                        'space':'Space','tab':'Tab','esc':'Escape','backspace':'Backspace','delete':'Delete','clear':'Clear',
                        'up':'ArrowUp','down':'ArrowDown','left':'ArrowLeft','right':'ArrowRight','pageup':'PageUp','pagedown':'PageDown',
                        'home':'Home','end':'End','insert':'Insert',
                    }

            map_dict2 = {'\+\]\[':'+','\]':']',
                    }


            # using lambda and regex functions to achieve task
            res = re.compile("|".join(map_dict1.keys())).sub(lambda ele: map_dict1[re.escape(ele.group(0))], test_str)
            res = re.compile("|".join(map_dict.keys())).sub(lambda ele: map_dict[re.escape(ele.group(0))], res) 
            res = re.compile("|".join(map_dict2.keys())).sub(lambda ele: map_dict2[re.escape(ele.group(0))], res)
            if res[-2:]=='+]': res=res[:-2]
            print('Truncate',res[-2:])
            # printing result
            print("The converted string : " + str(res))
            #res.split('][')
            import re
            pattern = r'[\[\],;|]'
            result = re.split(pattern, res)
            while("" in result): result.remove("")
            print(result)
            for i in result:
                print(i)
            return result

        for key in mapKeyCodes(codeValue):
            p.press(key)
    else:
        r.keyboard(codeValue)           # print('keyboard',prefix[1])

def _type(codeValue: str):              # type text at element
    identifier = codeValue.split(',',1)[0].strip()
    value = codeValue.split(',',1)[1].strip()
    if RPABROWSER == 1: 
        p.input(identifier,value)
    else:   
        r.type(identifier,value)


def _select(codeValue: str):
    key = codeValue.split(',',1)[0]
    value = codeValue.split(',',1)[1]
    #print("select", key, value)
    #logg('select:', select = key, option = value)
    if RPABROWSER == 1:
        p.select_option(key, value)
    else:
        r.select(key,value)             # print('select',prefix[1])
    #print("select done")

def _snap(codeValue):
    page = codeValue.split(',',1)[0]        
    if page=='':
        page = 'page'
    elif page =='log':
        snapImage()
        return
    saveFile = codeValue.split(',',1)[1]
    if saveFile=='': saveFile = './Output/Images/' + saveFile
    #logg('######### snap ###########', page = page, saveFile = saveFile, level = 'debug')
    if RPABROWSER == 1: 
        p.snap(path=saveFile, full_page=True)
    else:
        r.snap(page, saveFile)          

def _telegram(codeValue):
    id = codeValue.split(',',1)[0].strip()
    msg = codeValue.split(',',1)[1].strip()
    #flow_run_name = context.get_run_context().flow_run.dict()['name']  # error if this is not executed in a flow
    from config import flow_run_name 
    r.telegram(int(id),f"{flow_run_name}-{msg}")

def _download(codeValue):
    download_url = codeValue.split(',',1)[0].strip()
    filename_to_save = codeValue.split(',',1)[1].strip()
    if RPABROWSER == 1:
        p.download(download_url, filename_to_save)
    else:
        r.download(download_url, filename_to_save)

def selectTable(codeValue: str, df):
    # Returns a objTableSet (table object) with parameter which is either a table object in the list or a worksheet
    logger = get_run_logger()
    # table specified
    obj = codeValue
    #from openpyxl import load_workbook
    #wb = load_workbook(STARTFILE, read_only=True)   # open an Excel file and return a workbook
    if obj in df[(df.Type == 'table')]['Object'].dropna().values.tolist():
        #objTableSet = df[(df.Type == 'key') & ((df.Object == obj))]
        objTableSet = df[(df.Type == 'table') & ((df.Object == obj))]   # filter for specified object
        objTableSet = objTableSet.dropna(axis=1, how='all')  # drop all columns where values are nan
        #objTableSet = objTableSet.loc[:, ~objTableSet.columns.str.contains('^Unnamed')]  # remove unnamed columns
        #objTableSet.set_index("Type", inplace = True)
        objTableSet = objTableSet.reset_index(drop=True)    # remove index
        objTableSet = objTableSet.drop(['Type', 'Object'], axis=1)  # drop Type and Object column  
        #logger.info(objTableSet.head())
        objTableSet.columns = objTableSet.iloc[0]   # promote 1st row to header
        objTableSet = objTableSet[1:] #take the data less the header row
        #objTableSet = objTableSet.drop(columns=[0])

        #new_header = objTableSet.iloc[0] #grab the first row for the header
        #objTableSet = objTableSet[1:] #take the data less the header row
        #objTableSet.set_axis(new_header, axis=1, inplace=True)
        #objTableSet.columns = new_header #set the header row as the df header
        #logger.info(objTableSet.head())
        return True, objTableSet
    else:
        result, objTableSet = _isWorkSheetName(obj, excelfile=STARTFILE)
        return result, objTableSet

def _isWorkSheetName(obj, excelfile):
    #check if obj is a worksheet in excelfile and returns result of sheet as dataframe
    logger = get_run_logger()
    try:
        #elif obj in wb.sheetnames:
        sheet = obj
        #excelfile = STARTFILE
        objTableSet = pd.read_excel(excelfile, sheet_name=sheet)
        #logger.info('sheet exists')
        #logger.info(objTableSet)
        return True, objTableSet
    except ValueError as e:
        pass
    return False, None

def _test(codeValue, df, objVar):
    logger = get_run_logger()
    result, objTableSet = selectTable(codeValue, df)  # returns objTableSet from worksheet name
    if result: logger.info(objTableSet.head())

def _ifObjectExist(var:str):
    var_exists = var in locals() or var in globals()
    #try:
    #    exec(var)
    #except NameError:
    #    var_exists = False
    #else:
    #    var_exists = True
    print(f"====================================================== {var} {var_exists}")
    #print(locals())
    #print(globals())
    return var_exists

def _email(codeValue, df):
    import json
    logger = get_run_logger()
    #logger.info(f"Email codeValue: {codeValue}")

    # parse codeValue:
    try:
        emailObj = json.loads(codeValue)    # if codeValue is a Json object
        #logger.info(emailObj)
        #logger.info(emailObj['To'])

    except ValueError as e:
        #logger.info(f"ValueError")
        # not a json format - i.e. assume its a table
        # raise ValueError(f"Raise Error: incorrect json format, {codeValue}")
        if len(codeValue.split(',')) > 1:
            colsMapping = codeValue.split(',',1)[1].strip()
            codeValue = codeValue.split(',',1)[0].strip()
            #logger.info(f"colsMapping:{colsMapping}, codeValue: {codeValue}")
        result, objTableSet = selectTable(codeValue, df)
        #if _ifObjectExist('colsMapping'): # rename columns to names required by email send function
        if 'colsMapping' in locals(): #or var in globals()
            #logger.info(f"Rename cols: {json.loads(colsMapping)}")
            #logger.info(json.loads(colsMapping))
            objTableSet.rename(columns = json.loads(colsMapping), inplace = True)
        mailfieldlst = ['To', 'CC', 'Subject', 'Body', 'HTMLBody', 'Attachment', 'boolDisplay', 'boolRun', 'boolForce']
        objTableSet = objTableSet[objTableSet.columns.intersection(mailfieldlst)]  # select columns for mailing
        if result == True:  # is a table
            n = constants['iterationCount']
            #logger.info('iterationCount ' + str(n))
            objTableSet = objTableSet.iloc[n]  # fiter tableset for current iteration row
            #logger.info(objTableSet)
            #logger.info('columns')
            emailObj = json.loads(objTableSet.to_json(orient="columns"))
            logger.debug(f"{log_space}{emailObj}")
    #logger.info(type(emailObj)) # dictionary object

    # Send email
    #print('*******************************************************************')
    #logger.info(f"****** Send Email - emailObj:{emailObj}")
    import traceback
    try:
        To = emailObj['To']
        Subject = emailObj['Subject']
        #HTMLBody = emailObj['HTMLBody']
        #Attachment = emailObj['Attachment']
        #email_sender.send_email(boolDisplay=False, To=To, Subject=Subject, HTMLBody=HTMLBody, Attachment=Attachment)
        #logger.info(f"******** Email To:{To} Subject:{Subject}")

        boolRun = emailObj['boolRun'] if 'boolRun' in emailObj else True 
        boolDisplay = emailObj['boolDisplay'] if 'boolDisplay' in emailObj else False
        boolForce = emailObj['boolForce'] if 'boolForce' in emailObj else False        

        if boolRun: 
            #logger.info(emailObj)
            email_sender = EmailsSender()

            #from auto_utility_email import sentEmailSubjectList
            #logger.info(f"#####>>>> sentEmailSubList, {len(sentEmailSubjectList)}")
            import datetime
            cutOffDateTme=datetime.datetime.today().replace(hour=int(variables['sentEmailCheck_hour']), minute=int(variables['sentEmailCheck_min']), second=0, microsecond=0)
            #logger.debug(f"{log_space}{cutOffDateTme}  {variables['sentEmailCheck_hour']}  {variables['sentEmailCheck_min']}")            
            sentEmailSubjectList = email_sender.getSentEmailSubjectList(sentEmailSubjectList = [], cutOffDateTme=cutOffDateTme)
            if not Subject in sentEmailSubjectList or boolForce:
                email_sender.send_email(boolDisplay=boolDisplay, boolRun=boolRun, EmailObj = emailObj)
                logger.debug(f'{log_space}email SENT')
            else:
                logger.debug(f'{log_space}email NOT SENT - already sent')
        else:
            logger.debug(f'{log_space}boolRun is FALSE')
        #result = email_sender.wait_send_complete()
    except ValueError as e:
        logger.error('error --', e)
    except Exception as e:
        logger.error(traceback.format_exc())
        #logger.info('error --', e)
    #print('complete')

def _waitEmailComplete(codeValue, df):
    logger = get_run_logger()
    email_sender = EmailsSender()
    result = email_sender.wait_send_complete()
    #import datetime
    #today = datetime.datetime.today().replace(hour=0, minute=0, second=0, microsecond=0) #.strftime('%Y-%m-%d')

    #print(today.strftime('%d/%m/%Y %H:%M %p'))
    #result2 = email_sender.folderItemsList(ofolder=5,dateRange_StartOn=today)          #.sentFolderList()
    #print(result2)
    
    #from auto_utility_email import today, sentEmailSubjectList
    #global today
    #global sentEmailSubjectList
    #print(today.strftime('%d/%m/%Y %H:%M %p'), sentEmailSubjectList)
    #print(sentEmailSubjectList)
 
    #email_sender.refreshMail()
 
    logger.debug(f'{log_space}Email complete = ' + str(result))


def etest():    
    logger = get_run_logger()

    lst = codeValue.split(',')
    #code_dct = {lst[i]: lst[i + 1] for i in range(0, len(lst), 2)}
    code_dct = {lst[i].split('=')[0]: lst[i].split('=')[1] for i in range(0, len(lst))}           
    
    # Driver code
    #lst = ['a', 1, 'b', 2, 'c', 3]
    logger.info(lst)
    logger.info(code_dct)


    #logger.info(f"   email:{id} msg:{msg}")
    email_sender = EmailsSender()
    #attachment=[r"C:\Users\roh\OneDrive\Personal\TravelPlan\Raymond's KR and SG visit agenda_v3.xlsx"] #"Path to the attachment"
    #email_sender.send_email(To="tsoh20@gmail.com", Subject="Test", HTMLBody="test", attachment=attachment)
    #print('Email', email_sender.wait_send_complete())


#@task
def _click(code:str):
    try:
        int(len(code))
    except ValueError:
        # not int
        #logg('Not int - code is nan', level = 'error')
        return
    print('click', code)
    if RPABROWSER == 1:
        p.click(code)        
    else:    
        if not waitIdentifierExist(code, 30, 1):        # identifier, time_sec, interval
                #logg('      Time out - unable to process step', level = 'critical')           #constants['lastCodelist'] = codeList  
                raise CriticalAccessFailure("Exception critical failure")              
                return
        if code.lower().endswith(('.png', '.jpg', '.jpeg')): code = Path(IMAGE_DIR + '/' + code).absolute().__str__() 
        r.click(code)
    print('      ','click',code)
    #logg('click', code = code, level = 'debug')


